import openpyxl

def readUsernameData(rownum,columnnum):
    book = openpyxl.load_workbook("C:\\Users\\jkann\\Documents\\Python\\Datadriven.xlsx")
    #by defaut its give incative, hence active
    sheet = book.active
    #serach sheet loginpage wuthin workbook and zeroth attributre aaayi save
    if "LoginPage" in book.sheetnames: #sheetname attribute
        sheet = book.worksheets[0]  #loads workbook with help of openpyxl _ built in funtion
        #return using index 1,0
        #Username = sheet.cell(row=2, column=1).value
        username = sheet.cell(row=rownum, column=columnnum).value
    return username


username = readUsernameData(2,2)
print(username)

#"C:\Users\jkann\Documents\Python\TestData.xlsx"