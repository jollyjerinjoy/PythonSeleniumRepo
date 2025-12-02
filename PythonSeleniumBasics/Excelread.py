#If you DO want to pass row & column numbers as parameter
import openpyxl


def readUsernameData():
    book = openpyxl.load_workbook("C:\\Users\\jkann\\Documents\\Python\\Datadriven.xlsx")
    # by defaut its give incative, hence active
    sheet = book.active

    if "LoginPage" in book.sheetnames:  # sheetname attribute
        sheet = book.worksheets[0]  # loads workbook with help of openpyxl _ built in funtion
        username = sheet.cell(row=2, column=1).value

    return username


username = readUsernameData()
print(username)
