import openpyxl
def readUsernamePassword():
    book = openpyxl.load_workbook("C:\\Users\\jkann\\Documents\\Python\\Datadriven.xlsx")
    # Select sheet
    if "LoginPage" in book.sheetnames:
        sheet = book["LoginPage"]
    else:
        sheet = book.active
    credentials = []
    row = 2  # data starts from row 2
    # while :-Keep reading rows until we find an empty one.
    while True:

        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        if username is None:
            break
        credentials.append((username, password))
        row += 1
    return credentials
# Usage
for user, pwd in readUsernamePassword():
    print("Username:", user, " Password:", pwd)