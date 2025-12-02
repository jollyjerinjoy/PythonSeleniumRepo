import openpyxl
class ExcelUtility:
    def read_user_data(self,rownum,columnnum):
        book = openpyxl.load_workbook("C:\\Users\\jkann\\Documents\\Python\\TestData.xlsx")
        sheet = book.active

        if "LoginPage" in book.sheetnames: #sheetname attribute
            sheet = book.worksheets[0]
            username = sheet.cell(row=rownum, column=columnnum).value
        return username

